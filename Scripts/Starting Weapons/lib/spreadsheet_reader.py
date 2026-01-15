"""
Spreadsheet reading functions for Race and Class Masks.xlsx
"""

import os
import openpyxl


def get_spreadsheet_path():
    """Get the path to the Race and Class Masks.xlsx file."""
    lib_dir = os.path.dirname(os.path.abspath(__file__))
    script_dir = os.path.dirname(lib_dir)
    return os.path.join(script_dir, 'Race and Class Masks.xlsx')


def read_weapon_skills_from_spreadsheet():
    """
    Read weapon skills from the 'Race Class Starting Skills' sheet.

    Returns:
        dict: {skill_id: {'name': str, 'classes': [{'name': str, 'class_mask': int, 'race_mask': int}]}}
    """
    spreadsheet_path = get_spreadsheet_path()

    print("=" * 80)
    print("READING WEAPON SKILLS FROM SPREADSHEET")
    print("=" * 80)
    print()

    if not os.path.exists(spreadsheet_path):
        print(f"ERROR: Spreadsheet not found at {spreadsheet_path}")
        return None

    wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)
    ws = wb['Race Class Starting Skills']

    weapon_skills = {}

    # Find all weapon skill tables
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row_idx, 1).value
        if cell_value and isinstance(cell_value, str) and 'Class Race - Starting Skill' in cell_value:
            parts = cell_value.split(' - ')
            if len(parts) >= 4:
                weapon_name = parts[2].strip()
                skill_id = int(parts[3].strip())
                mask_row = row_idx + 15

                class_names = ['Warrior', 'Paladin', 'Hunter', 'Rogue', 'Priest', 'Death Knight', 'Shaman', 'Mage', 'Warlock', 'Druid']
                class_masks = [1, 2, 4, 8, 16, 32, 64, 128, 256, 1024]

                if skill_id not in weapon_skills:
                    weapon_skills[skill_id] = {'name': weapon_name, 'classes': []}

                for col_idx, (class_name, class_mask) in enumerate(zip(class_names, class_masks), start=2):
                    race_mask = ws.cell(mask_row, col_idx).value
                    if race_mask and race_mask != 0:
                        weapon_skills[skill_id]['classes'].append({
                            'name': class_name,
                            'class_mask': class_mask,
                            'race_mask': race_mask
                        })

    print(f"Loaded {len(weapon_skills)} weapon skills from spreadsheet")
    print()

    return weapon_skills


def read_weapon_class_masks():
    """
    Read weapon class masks from the 'Weapon Class Mask' sheet.

    Returns:
        dict: {skill_id: {'name': str, 'class_mask': int}}
    """
    spreadsheet_path = get_spreadsheet_path()

    if not os.path.exists(spreadsheet_path):
        print(f"ERROR: Spreadsheet not found at {spreadsheet_path}")
        return {}

    wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)
    ws = wb['Weapon Class Mask']

    weapon_class_data = {}

    # Row 2 has class masks (columns C-K)
    class_masks = []
    for col_idx in range(3, 12):  # C=3 to K=11
        mask_val = ws.cell(2, col_idx).value
        class_masks.append(int(mask_val) if mask_val else 0)

    # Row 3 has class names
    class_names = []
    for col_idx in range(3, 12):
        name = ws.cell(3, col_idx).value
        class_names.append(name.strip() if name else "")

    # Rows 4+ have weapon data
    for row_idx in range(4, ws.max_row + 1):
        weapon_name = ws.cell(row_idx, 1).value
        if not weapon_name:
            break

        skill_id = ws.cell(row_idx, 2).value
        if not skill_id:
            continue

        skill_id = int(skill_id)

        # Calculate class mask by OR-ing masks where class can learn
        combined_class_mask = 0
        for col_idx, (class_mask, class_name) in enumerate(zip(class_masks, class_names), start=3):
            can_learn = ws.cell(row_idx, col_idx).value
            if can_learn and str(can_learn).lower() in ['true', '1', 'yes']:
                combined_class_mask |= class_mask

        weapon_class_data[skill_id] = {
            'name': weapon_name,
            'class_mask': combined_class_mask
        }

    return weapon_class_data
