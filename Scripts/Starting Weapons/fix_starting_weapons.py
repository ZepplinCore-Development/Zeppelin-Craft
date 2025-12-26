#!/usr/bin/env python3
"""
Consolidated starting gear validation and fix generator.

This script:
1. Reads weapon skills from Race and Class Masks.xlsx
2. Generates skillraceclassinfo.sql (makes skills learnable)
3. Generates starting_weapon_skills.sql (starting skills at creation)
4. Validates starter weapon items against item_template
5. Checks all race/class charstartoutfit weapons against available skills
6. Generates starting_weapons.sql with fixes and ammo
"""

import pymysql
import sys
import openpyxl
import os

# Database connections
DBC_CONFIG = {
    'host': '192.168.0.55',
    'port': 3306,
    'user': 'spell-editor',
    'passwd': 'HW8Y%L6&f0ePJO',
    'db': 'dbc'
}

ACORE_CONFIG = {
    'host': '192.168.0.55',
    'port': 3306,
    'user': 'acore',
    'passwd': 'acore',
    'db': 'acore_world'
}

RACE_NAMES = {
    1: "Human", 
    2: "Orc", 
    3: "Dwarf", 
    4: "Night Elf", 
    5: "Undead",
    6: "Tauren", 
    7: "Gnome", 
    8: "Troll", 
    9: "Goblin", 
    10: "Blood Elf",
    11: "Draenei", 
    12: "Worgen"
}

CLASS_NAMES = {
    1: "Warrior", 
    2: "Paladin", 
    3: "Hunter", 
    4: "Rogue", 
    5: "Priest",
    6: "Death Knight", 
    7: "Shaman", 
    8: "Mage", 
    9: "Warlock", 
    10: "Monk",
    11: "Druid"
}

# Weapon subclass to skill mapping
WEAPON_SKILLS = {
    0: (44, "One-Handed Axes"),
    1: (172, "Two-Handed Axes"),
    4: (54, "One-Handed Maces"),
    5: (160, "Two-Handed Maces"),
    7: (43, "One-Handed Swords"),
    8: (55, "Two-Handed Swords"),
    10: (136, "Staves"),
    19: (228, "Wands"),          # Fixed: Wands are subclass 19, not 13
    15: (173, "Daggers"),
    2: (45, "Bows"),
    3: (46, "Guns"),
    18: (226, "Crossbows"),  # Crossbows are skill 226, not 47
    16: (176, "Thrown"),
}

# Candidate starter weapons by skill ID (to be validated)
STARTER_WEAPON_CANDIDATES = {
    44: [37],                 # 1H Axes - Worn Axe
    172: [12282],             # 2H Axes - Worn Battleaxe
    54: [36],                 # 1H Maces - Worn Mace
    160: [2361],              # 2H Maces - Battleworn Hammer
    43: [25],                 # 1H Swords - Worn Shortsword
    55: [49778, 23346],       # 2H Swords - Worn Greatsword, Battleworn Claymore
    136: [35, 3661, 20978],   # Staves - Bent Staff, Handcrafted Staff, Apprentice's Staff
    173: [2092],              # Daggers - Worn Dagger
    228: [4902],              # Wands - Apprentice Wand
    45: [2504],               # Bows - Worn Shortbow
    46: [2508],               # Guns - Old Blunderbuss
    226: [23347],             # Crossbows - Weathered Crossbow
    176: [25861],             # Thrown - Crude Throwing Axe
}

# Starter ammo for ranged weapons (invType 24)
STARTER_AMMO = {
    45: 2512,    # Bows -> Rough Arrow
    46: 2516,    # Guns -> Light Shot
    226: 2512,   # Crossbows -> Rough Arrow
}

# Weapon skill to invType mapping
SKILL_TO_INVTYPE = {
    43: 13,   # Swords -> One-Hand
    44: 13,   # Axes -> One-Hand
    54: 13,   # Maces -> One-Hand
    173: 13,  # Daggers -> One-Hand
    55: 17,   # Two-Handed Swords -> Two-Hand
    160: 17,  # Two-Handed Maces -> Two-Hand
    172: 17,  # Two-Handed Axes -> Two-Hand
    136: 17,  # Staves -> Two-Hand
    229: 17,  # Polearms -> Two-Hand
    45: 15,   # Bows -> Ranged
    46: 15,   # Guns -> Ranged
    226: 15,  # Crossbows -> Ranged
    176: 26,  # Thrown -> Ranged Right
    228: 26,  # Wands -> Ranged Right
    162: 13,  # Unarmed -> One-Hand (fist)
    433: 14,  # Shield -> Shield
    473: 13,  # Fist Weapons -> One-Hand
}

# Path to spreadsheet (relative to this script)
# Script is in: Zeppelin-Craft/Scripts/Item Scripts/Starting Items/
# Spreadsheet is in: Zeppelin-Craft/Scripts/Item Scripts/Starting Items/ (same folder)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(SCRIPT_DIR, 'Race and Class Masks.xlsx')

def read_weapon_skills_from_spreadsheet():
    """Read weapon skills from the Excel spreadsheet."""
    print("=" * 80)
    print("READING WEAPON SKILLS FROM SPREADSHEET")
    print("=" * 80)
    print()

    if not os.path.exists(SPREADSHEET_PATH):
        print(f"ERROR: Spreadsheet not found at {SPREADSHEET_PATH}")
        return None

    wb = openpyxl.load_workbook(SPREADSHEET_PATH, data_only=True)
    ws = wb['Race Class Starting Skills']

    weapon_skills = {}

    # Find all weapon skill tables
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row_idx, 1).value
        if cell_value and isinstance(cell_value, str) and 'Class Race - Starting Skill' in cell_value:
            parts = cell_value.split(' - ')
            if len(parts) >= 4:
                weapon_name = parts[2].strip()
                skill_id = int(parts[3].strip())
                mask_row = row_idx + 15

                class_names = ['Warrior', 'Paladin', 'Hunter', 'Rogue', 'Priest', 'Death Knight', 'Shaman', 'Mage', 'Warlock', 'Druid']
                class_masks = [1, 2, 4, 8, 16, 32, 64, 128, 256, 1024]

                if skill_id not in weapon_skills:
                    weapon_skills[skill_id] = {'name': weapon_name, 'classes': []}

                for col_idx, (class_name, class_mask) in enumerate(zip(class_names, class_masks), start=2):
                    race_mask = ws.cell(mask_row, col_idx).value
                    if race_mask and race_mask != 0:
                        weapon_skills[skill_id]['classes'].append({
                            'name': class_name,
                            'class_mask': class_mask,
                            'race_mask': race_mask
                        })

    print(f"Loaded {len(weapon_skills)} weapon skills from spreadsheet")
    print()

    return weapon_skills

def read_weapon_class_masks():
    """Read weapon class masks from the 'Weapon Class Mask' sheet."""
    if not os.path.exists(SPREADSHEET_PATH):
        print(f"ERROR: Spreadsheet not found at {SPREADSHEET_PATH}")
        return {}

    wb = openpyxl.load_workbook(SPREADSHEET_PATH, data_only=True)
    ws = wb['Weapon Class Mask']

    weapon_class_data = {}

    # Row 2 has class masks (columns C-K)
    class_masks = []
    for col_idx in range(3, 12):  # C=3 to K=11
        mask_val = ws.cell(2, col_idx).value
        class_masks.append(int(mask_val) if mask_val else 0)

    # Row 3 has class names
    class_names = []
    for col_idx in range(3, 12):
        name = ws.cell(3, col_idx).value
        class_names.append(name.strip() if name else "")

    # Rows 4+ have weapon data
    for row_idx in range(4, ws.max_row + 1):
        weapon_name = ws.cell(row_idx, 1).value
        if not weapon_name:
            break

        skill_id = ws.cell(row_idx, 2).value
        if not skill_id:
            continue

        skill_id = int(skill_id)

        # Calculate class mask by OR-ing masks where class can learn
        combined_class_mask = 0
        for col_idx, (class_mask, class_name) in enumerate(zip(class_masks, class_names), start=3):
            can_learn = ws.cell(row_idx, col_idx).value
            if can_learn and str(can_learn).lower() in ['true', '1', 'yes']:
                combined_class_mask |= class_mask

        weapon_class_data[skill_id] = {
            'name': weapon_name,
            'class_mask': combined_class_mask
        }

    return weapon_class_data

def generate_skillraceclassinfo_sql():
    """Generate skillraceclassinfo.sql to make skills learnable by class."""
    print("=" * 80)
    print("GENERATING SKILL RACE CLASS INFO SQL")
    print("=" * 80)
    print()

    # Read weapon class masks from spreadsheet
    weapon_class_data = read_weapon_class_masks()

    sql_lines = []

    # Header
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("-- SKILL RACE CLASS INFO - WEAPON SKILL AVAILABILITY")
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("-- This file controls which weapon skills are LEARNABLE by CLASS.")
    sql_lines.append("-- All races (raceMask=4095) can learn if their class allows it.")
    sql_lines.append("-- Must be applied BEFORE starting_weapon_skills.sql")
    sql_lines.append("--")
    sql_lines.append("-- AUTO-GENERATED from Race and Class Masks.xlsx (Weapon Class Mask sheet)")
    sql_lines.append("-- DO NOT EDIT MANUALLY - Changes will be overwritten")
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("")

    # Delete existing weapon skill entries
    all_skill_ids = sorted(weapon_class_data.keys())
    skill_id_list = ', '.join(str(sid) for sid in all_skill_ids)
    sql_lines.append("-- Delete existing weapon skill entries")
    sql_lines.append(f"DELETE FROM `skillraceclassinfo` WHERE `SkillLineDbcRecord` IN ({skill_id_list});")
    sql_lines.append("")

    # Insert entries for each skill with explicit IDs
    # Start from ID 1000 to avoid conflicts with existing entries
    entry_id = 1000
    for skill_id in all_skill_ids:
        data = weapon_class_data[skill_id]
        sql_lines.append(f"-- {data['name']} (skill {skill_id}) - classMask {data['class_mask']}")

        # Insert with explicit ID, raceMask=4095 (all races), flags=128, minLevel=0, skillTierID=0, skillCostIndex=0
        sql_lines.append(
            f"INSERT INTO `skillraceclassinfo` "
            f"(`Id`, `SkillLineDbcRecord`, `RaceMask`, `ClassMask`, `Flags`, `MinLevel`, `SkillTierId`, `SkillCostIndex`) "
            f"VALUES ({entry_id}, {skill_id}, 4095, {data['class_mask']}, 128, 0, 0, 0);"
        )
        sql_lines.append("")
        entry_id += 1

    sql_content = '\n'.join(sql_lines)

    # Write to file (same directory as script)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, 'skillraceclassinfo.sql')
    with open(output_path, 'w') as f:
        f.write(sql_content)

    print(f"✓ Generated {output_path}")
    print(f"  {len(all_skill_ids)} weapon skills")
    print(f"  All races (raceMask=4095) with class-specific masks")
    print()

def generate_starting_skills_sql(weapon_skills):
    """Generate starting_weapon_skills.sql from weapon skills data."""
    print("=" * 80)
    print("GENERATING WEAPON SKILLS SQL")
    print("=" * 80)
    print()

    sql_lines = []

    # Header
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("-- STARTING WEAPON SKILLS")
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("-- This file controls which races can use which weapon types at character")
    sql_lines.append("-- creation. All weapon skill changes should be made in this file ONLY.")
    sql_lines.append("--")
    sql_lines.append("-- AUTO-GENERATED from Race and Class Masks.xlsx")
    sql_lines.append("-- DO NOT EDIT MANUALLY - Changes will be overwritten")
    sql_lines.append("--")
    sql_lines.append("-- RACE MASKS (add values for multiple races):")
    sql_lines.append("--   1 = Human       2 = Orc         4 = Dwarf       8 = Night Elf")
    sql_lines.append("--  16 = Undead     32 = Tauren     64 = Gnome     128 = Troll")
    sql_lines.append("-- 256 = Goblin    512 = Blood Elf  1024 = Draenei  2048 = Worgen")
    sql_lines.append("--   0 = None      4095 = All Races")
    sql_lines.append("--")
    sql_lines.append("-- CLASS MASKS (add values for multiple classes):")
    sql_lines.append("--   1 = Warrior     2 = Paladin      4 = Hunter      8 = Rogue     16 = Priest")
    sql_lines.append("--  32 = DK         64 = Shaman     128 = Mage      256 = Warlock  512 = (unused)")
    sql_lines.append("-- 1024 = Druid")
    sql_lines.append("--   0 = All       2047 = All Classes")
    sql_lines.append("-- " + "=" * 76)
    sql_lines.append("")

    # Delete existing weapon skills
    sql_lines.append("-- Delete existing weapon skills")
    all_skill_ids = sorted(weapon_skills.keys())
    skill_id_list = ', '.join(str(sid) for sid in all_skill_ids)
    sql_lines.append(f"DELETE FROM `playercreateinfo_skills` WHERE `skill` IN ({skill_id_list});")
    sql_lines.append("")

    # Group by weapon category
    weapon_categories = {
        'ONE-HANDED WEAPONS': [43, 44, 54, 173],
        'TWO-HANDED WEAPONS': [55, 160, 172, 136, 229],
        'RANGED WEAPONS': [45, 46, 47, 176, 228],
        'OTHER': [162, 433, 473]
    }

    for category, skill_ids in weapon_categories.items():
        category_has_skills = any(sid in weapon_skills for sid in skill_ids)
        if not category_has_skills:
            continue

        sql_lines.append("-- " + "=" * 76)
        sql_lines.append(f"-- {category}")
        sql_lines.append("-- " + "=" * 76)
        sql_lines.append("")

        for skill_id in skill_ids:
            if skill_id not in weapon_skills:
                continue

            data = weapon_skills[skill_id]
            sql_lines.append(f"-- {data['name']} (skill {skill_id})")

            for cls in data['classes']:
                comment = f"{cls['name']} - {data['name']}"
                # Escape single quotes for SQL
                comment_escaped = comment.replace("'", "''")
                sql_lines.append(f"INSERT INTO `playercreateinfo_skills` (`raceMask`, `classMask`, `skill`, `rank`, `comment`) VALUES ({cls['race_mask']}, {cls['class_mask']}, {skill_id}, 0, '{comment_escaped}');")

            sql_lines.append("")

    sql_content = '\n'.join(sql_lines)

    # Write to file (same directory as script)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, 'starting_weapon_skills.sql')
    with open(output_path, 'w') as f:
        f.write(sql_content)

    print(f"✓ Generated {output_path}")
    print(f"  {len(all_skill_ids)} weapon skills")
    print(f"  {sum(len(data['classes']) for data in weapon_skills.values())} total class entries")
    print()

    return weapon_skills

def validate_starter_weapons(cursor):
    """Validate starter weapon candidates against item_template."""
    print("=" * 80)
    print("VALIDATING STARTER WEAPON ITEMS")
    print("=" * 80)
    print()

    validated_weapons = {}

    for skill_id, candidate_items in STARTER_WEAPON_CANDIDATES.items():
        skill_name = next((name for subclass, (sid, name) in WEAPON_SKILLS.items() if sid == skill_id), f"Skill {skill_id}")

        if not candidate_items:
            # Special case: find wands
            if skill_id == 228:
                cursor.execute("""
                    SELECT entry, name, subclass
                    FROM item_template
                    WHERE class = 2 AND subclass = 19
                    AND RequiredLevel <= 1
                    ORDER BY entry
                    LIMIT 5
                """)
                wands = cursor.fetchall()
                if wands:
                    print(f"{skill_name} (skill {skill_id}):")
                    for entry, name, subclass in wands:
                        print(f"  Found: {entry} - {name} (subclass {subclass})")
                    candidate_items = [wands[0][0]]  # Use first wand
                else:
                    print(f"{skill_name} (skill {skill_id}): No suitable starter wands found!")
                    continue

        # Verify each candidate
        print(f"{skill_name} (skill {skill_id}):")
        for item_id in candidate_items:
            cursor.execute("""
                SELECT entry, name, class, subclass
                FROM item_template
                WHERE entry = %s
            """, (item_id,))
            item = cursor.fetchone()

            if not item:
                print(f"  ✗ {item_id}: Item not found in database!")
                continue

            entry, name, item_class, subclass = item

            # Check if it's a weapon (class 2)
            if item_class != 2:
                print(f"  ✗ {entry} ({name}): Not a weapon (class {item_class})")
                continue

            # Check if subclass matches the skill
            expected_subclass = next((sc for sc, (sid, sname) in WEAPON_SKILLS.items() if sid == skill_id), None)

            if expected_subclass is not None and subclass == expected_subclass:
                print(f"  ✓ {entry} ({name}): Valid starter weapon")
                if skill_id not in validated_weapons:
                    validated_weapons[skill_id] = entry
            else:
                actual_skill = WEAPON_SKILLS.get(subclass, (None, "Unknown"))
                print(f"  ✗ {entry} ({name}): Wrong subclass {subclass} ({actual_skill[1]}) - expected {expected_subclass}")

        print()

    # Report summary
    missing = []
    for skill_id in STARTER_WEAPON_CANDIDATES.keys():
        if skill_id not in validated_weapons:
            skill_name = next((name for subclass, (sid, name) in WEAPON_SKILLS.items() if sid == skill_id), f"Skill {skill_id}")
            missing.append(f"{skill_name} (skill {skill_id})")

    if missing:
        print("=" * 80)
        print("MISSING VALID STARTER WEAPONS:")
        print("=" * 80)
        for item in missing:
            print(f"  - {item}")
        print()

    return validated_weapons

def get_race_mask_bit(race_id):
    return 1 << (race_id - 1)

def get_class_mask_bit(class_id):
    return 1 << (class_id - 1)

def is_valid_race_class_combo(cursor, race_id, class_id):
    """Check if a race/class combination exists in playercreateinfo."""
    query = """
        SELECT COUNT(*) FROM playercreateinfo
        WHERE race = %s AND class = %s
    """
    cursor.execute(query, (race_id, class_id))
    result = cursor.fetchone()
    return result[0] > 0

def get_available_weapon_skills(weapon_skills, race_id, class_id):
    """Get all weapon skills available to a race/class combination from spreadsheet data."""
    race_bit = get_race_mask_bit(race_id)
    class_bit = get_class_mask_bit(class_id)

    available_skills = []

    # Check each weapon skill
    for skill_id, data in weapon_skills.items():
        # Check each class entry for this skill
        for cls in data['classes']:
            # Check if this class mask matches
            if cls['class_mask'] & class_bit:
                # Check if this race mask matches
                race_mask = cls['race_mask']
                if race_mask == 4095 or (race_mask & race_bit):
                    available_skills.append((skill_id, f"{cls['name']} - {data['name']}"))
                    break  # Found a match for this skill, no need to check other class entries

    return available_skills

def main():
    validate_only = '--validate-only' in sys.argv

    print("=" * 80)
    print("STARTING GEAR FIX GENERATOR")
    print("=" * 80)
    print()

    # Step 1: Read weapon skills from spreadsheet
    weapon_skills = read_weapon_skills_from_spreadsheet()
    if not weapon_skills:
        print("ERROR: Failed to read weapon skills from spreadsheet!")
        return 1

    # Step 2: Generate skillraceclassinfo.sql (must be applied FIRST)
    generate_skillraceclassinfo_sql()

    # Step 3: Generate starting_weapon_skills.sql
    generate_starting_skills_sql(weapon_skills)

    # Step 4: Connect to databases
    try:
        dbc_conn = pymysql.connect(**DBC_CONFIG)
        acore_conn = pymysql.connect(**ACORE_CONFIG)
    except pymysql.Error as e:
        print(f"Database connection error: {e}")
        return 1

    dbc_cursor = dbc_conn.cursor()
    acore_cursor = acore_conn.cursor()

    # Step 3: Validate starter weapons
    starter_weapons = validate_starter_weapons(acore_cursor)

    if not starter_weapons:
        print("ERROR: No valid starter weapons found! Cannot generate fixes.")
        return 1

    # Step 4: Validate starting gear
    print("=" * 80)
    print("VALIDATING STARTING GEAR")
    print("=" * 80)
    print()

    mismatches = []
    weapon_additions = []  # Track additional weapons needed (melee/ranged)
    skipped_invalid = 0

    # Get all outfits with weapons
    query = """
        SELECT o.ID, o.race, o.class, o.gender,
               o.itemId1, o.itemId2, o.itemId3, o.itemId4, o.itemId5,
               o.itemId6, o.itemId7, o.itemId8, o.itemId9, o.itemId10,
               o.itemId11, o.itemId12, o.itemId13, o.itemId14, o.itemId15,
               o.itemId16, o.itemId17, o.itemId18, o.itemId19, o.itemId20,
               o.itemId21, o.itemId22, o.itemId23, o.itemId24,
               o.invType1, o.invType2, o.invType3, o.invType4, o.invType5,
               o.invType6, o.invType7, o.invType8, o.invType9, o.invType10,
               o.invType11, o.invType12, o.invType13, o.invType14, o.invType15,
               o.invType16, o.invType17, o.invType18, o.invType19, o.invType20,
               o.invType21, o.invType22, o.invType23, o.invType24
        FROM charstartoutfit o
        WHERE o.race BETWEEN 1 AND 12 AND o.class BETWEEN 1 AND 11
    """

    dbc_cursor.execute(query)
    outfits = dbc_cursor.fetchall()

    for outfit in outfits:
        outfit_id = outfit[0]
        race_id = outfit[1]
        class_id = outfit[2]
        gender = outfit[3]
        item_ids = outfit[4:28]  # itemId1-24 (24 slots)
        inv_types = outfit[28:52]  # invType1-24 (24 slots)

        race_name = RACE_NAMES.get(race_id, f"Race{race_id}")
        class_name = CLASS_NAMES.get(class_id, f"Class{class_id}")
        gender_name = "Male" if gender == 0 else "Female"

        # Skip invalid race/class combinations
        if not is_valid_race_class_combo(acore_cursor, race_id, class_id):
            skipped_invalid += 1
            continue

        # Get available weapon skills from spreadsheet data
        available_skills = get_available_weapon_skills(weapon_skills, race_id, class_id)
        available_skill_ids = [skill_id for skill_id, _ in available_skills]

        # Track what weapons this outfit will have after fixes
        outfit_weapons = {}  # slot_idx -> (skill_id, item_id)
        weapon_inv_types = [13, 15, 17, 21, 22, 26]

        # First pass: Check each weapon slot for mismatches
        for slot_idx, (item_id, inv_type) in enumerate(zip(item_ids, inv_types), 1):
            if not item_id or item_id <= 0 or item_id == 6948:
                continue

            if inv_type not in weapon_inv_types:
                continue

            # Get item details
            acore_cursor.execute(
                "SELECT entry, name, class, subclass FROM item_template WHERE entry = %s",
                (item_id,)
            )
            item = acore_cursor.fetchone()

            if not item or item[2] != 2:  # Not a weapon
                continue

            item_name = item[1]
            subclass = item[3]

            # Check if they have the required skill
            if subclass not in WEAPON_SKILLS:
                continue

            required_skill_id, weapon_type = WEAPON_SKILLS[subclass]

            if required_skill_id not in available_skill_ids:
                # Found a mismatch - find replacement
                replacement_item = None
                replacement_skill = None

                # Prioritize similar weapon types with fallbacks
                if required_skill_id in [43, 44, 54]:  # 1H melee
                    for skill in [54, 44, 43, 173]:
                        if skill in available_skill_ids and skill in starter_weapons:
                            replacement_skill = skill
                            replacement_item = starter_weapons[skill]
                            break
                elif required_skill_id in [55, 160, 172]:  # 2H melee
                    for skill in [160, 172, 55, 54, 44, 43, 173]:
                        if skill in available_skill_ids and skill in starter_weapons:
                            replacement_skill = skill
                            replacement_item = starter_weapons[skill]
                            break
                elif required_skill_id == 136:  # Staves - casters get wand instead
                    if 228 in available_skill_ids and 228 in starter_weapons:
                        replacement_skill = 228
                        replacement_item = starter_weapons[228]
                elif required_skill_id in [45, 46, 47]:  # Ranged
                    for skill in [46, 45, 47, 176]:
                        if skill in available_skill_ids and skill in starter_weapons:
                            replacement_skill = skill
                            replacement_item = starter_weapons[skill]
                            break
                elif required_skill_id == 173:  # Daggers
                    for skill in [173, 54, 44, 43]:
                        if skill in available_skill_ids and skill in starter_weapons:
                            replacement_skill = skill
                            replacement_item = starter_weapons[skill]
                            break
                elif required_skill_id == 228:  # Wands
                    if 228 in available_skill_ids and 228 in starter_weapons:
                        replacement_skill = 228
                        replacement_item = starter_weapons[228]

                mismatches.append({
                    'outfit_id': outfit_id,
                    'race': race_name,
                    'class': class_name,
                    'gender': gender_name,
                    'slot': slot_idx,
                    'current_item': item_id,
                    'current_name': item_name,
                    'current_type': weapon_type,
                    'required_skill': required_skill_id,
                    'replacement_item': replacement_item,
                    'replacement_skill': replacement_skill,
                    'available_skills': available_skill_ids
                })

                # Track the replacement
                if replacement_skill:
                    outfit_weapons[slot_idx] = (replacement_skill, replacement_item)
            else:
                # Valid weapon - track it
                outfit_weapons[slot_idx] = (required_skill_id, item_id)

        # Second pass: Generic rule - ensure they have weapons for their available skills
        # Melee weapon skills
        melee_skills = [43, 44, 54, 55, 136, 160, 172, 173]  # All melee weapon types
        # Ranged weapon skills
        ranged_skills = [45, 46, 226, 176, 228]  # Bows, Guns, Crossbows, Thrown, Wands

        # Check what types of skills they have
        has_melee_skill = any(skill in available_skill_ids for skill in melee_skills)
        has_ranged_skill = any(skill in available_skill_ids for skill in ranged_skills)

        # Check what types of weapons they currently have (after replacements)
        has_melee_weapon = any(skill_id in melee_skills for skill_id, _ in outfit_weapons.values())
        has_ranged_weapon = any(skill_id in ranged_skills for skill_id, _ in outfit_weapons.values())

        # Find first empty slot for additions
        empty_slots = [i for i in range(1, 25) if item_ids[i-1] <= 0 or item_ids[i-1] == 0]

        # If they have melee skill but no melee weapon, add one
        if has_melee_skill and not has_melee_weapon:
            # Find the first available melee skill they have
            for skill in melee_skills:
                if skill in available_skill_ids and skill in starter_weapons:
                    if empty_slots:
                        slot = empty_slots.pop(0)
                        skill_name = next((name for sid, name in WEAPON_SKILLS.values() if sid == skill), f"Skill {skill}")
                        weapon_additions.append({
                            'outfit_id': outfit_id,
                            'race': race_name,
                            'class': class_name,
                            'gender': gender_name,
                            'slot': slot,
                            'add_item': starter_weapons[skill],
                            'add_skill': skill,
                            'reason': f'Has melee skill, needs melee weapon'
                        })
                        break

        # If they have ranged skill but no ranged weapon, add one
        if has_ranged_skill and not has_ranged_weapon:
            # Find the first available ranged skill they have
            for skill in ranged_skills:
                if skill in available_skill_ids and skill in starter_weapons:
                    if empty_slots:
                        slot = empty_slots.pop(0)
                        skill_name = next((name for sid, name in WEAPON_SKILLS.values() if sid == skill), f"Skill {skill}")
                        weapon_additions.append({
                            'outfit_id': outfit_id,
                            'race': race_name,
                            'class': class_name,
                            'gender': gender_name,
                            'slot': slot,
                            'add_item': starter_weapons[skill],
                            'add_skill': skill,
                            'reason': f'Has ranged skill, needs ranged weapon'
                        })

                        # If this weapon needs ammo (bows, guns, crossbows), add ammo too
                        if skill in STARTER_AMMO and empty_slots:
                            ammo_slot = empty_slots.pop(0)
                            ammo_item = STARTER_AMMO[skill]
                            weapon_additions.append({
                                'outfit_id': outfit_id,
                                'race': race_name,
                                'class': class_name,
                                'gender': gender_name,
                                'slot': ammo_slot,
                                'add_item': ammo_item,
                                'add_skill': None,  # Ammo doesn't have a skill
                                'reason': f'Ammo for {skill_name}',
                                'is_ammo': True
                            })
                        break

    # Report results
    weapons_only = [a for a in weapon_additions if not a.get('is_ammo')]
    ammo_only = [a for a in weapon_additions if a.get('is_ammo')]
    print(f"Skipped {skipped_invalid} invalid race/class combinations")
    print(f"Found {len(mismatches)} weapon mismatches in valid combinations")
    print(f"Found {len(weapons_only)} weapon additions needed (melee/ranged)")
    print(f"Found {len(ammo_only)} ammo additions needed (for bows/guns/crossbows)")
    print()

    if validate_only:
        # Just print validation results
        if mismatches or weapon_additions:
            print("=" * 80)
            print("ISSUES FOUND")
            print("=" * 80)
            print()
            if mismatches:
                print("MISMATCHES:")
                for mismatch in mismatches:
                    print(f"{mismatch['race']} {mismatch['class']} ({mismatch['gender']}):")
                    print(f"  Slot {mismatch['slot']}: {mismatch['current_item']} ({mismatch['current_name']}) - {mismatch['current_type']}")
                    print(f"  Problem: Requires skill {mismatch['required_skill']} (not available)")
                    print()
            if weapon_additions:
                print("WEAPON ADDITIONS:")
                for addition in weapon_additions:
                    skill_name = next((name for sid, name in WEAPON_SKILLS.values() if sid == addition['add_skill']), "Unknown")
                    print(f"{addition['race']} {addition['class']} ({addition['gender']}):")
                    print(f"  Need to add: {skill_name} ({addition['reason']})")
                    print()
        else:
            print("✓ NO ISSUES FOUND!")
    else:
        # First, fix any items that have itemId but missing invType
        print("=" * 80)
        print("FIXING MISSING INVTYPE VALUES")
        print("=" * 80)
        print()

        invtype_fixes = []
        dbc_cursor.execute("""
            SELECT o.ID, o.race, o.class, o.gender,
                   o.itemId1, o.itemId2, o.itemId3, o.itemId4, o.itemId5,
                   o.itemId6, o.itemId7, o.itemId8, o.itemId9, o.itemId10,
                   o.itemId11, o.itemId12, o.itemId13, o.itemId14, o.itemId15,
                   o.itemId16, o.itemId17, o.itemId18, o.itemId19, o.itemId20,
                   o.itemId21, o.itemId22, o.itemId23, o.itemId24,
                   o.invType1, o.invType2, o.invType3, o.invType4, o.invType5,
                   o.invType6, o.invType7, o.invType8, o.invType9, o.invType10,
                   o.invType11, o.invType12, o.invType13, o.invType14, o.invType15,
                   o.invType16, o.invType17, o.invType18, o.invType19, o.invType20,
                   o.invType21, o.invType22, o.invType23, o.invType24
            FROM charstartoutfit o
            WHERE o.race BETWEEN 1 AND 12 AND o.class BETWEEN 1 AND 11
        """)
        current_outfits = dbc_cursor.fetchall()

        for outfit in current_outfits:
            outfit_id = outfit[0]
            race_id = outfit[1]
            class_id = outfit[2]
            gender = outfit[3]
            item_ids = outfit[4:28]  # itemId1-24 (24 slots)
            inv_types = outfit[28:52]  # invType1-24 (24 slots)

            race_name = RACE_NAMES.get(race_id, f"Race{race_id}")
            class_name = CLASS_NAMES.get(class_id, f"Class{class_id}")
            gender_name = "Male" if gender == 0 else "Female"

            for slot_idx, (item_id, inv_type) in enumerate(zip(item_ids, inv_types), 1):
                # Skip empty slots and hearthstone
                if not item_id or item_id <= 0 or item_id == 6948:
                    continue

                # Look up item in item_template to determine correct invType
                acore_cursor.execute(
                    "SELECT class, subclass, name, InventoryType FROM item_template WHERE entry = %s",
                    (item_id,)
                )
                result = acore_cursor.fetchone()

                if result:
                    item_class, subclass, item_name, correct_inv_type = result

                    # Weapons (class 2) - use skill-based invType mapping
                    if item_class == 2 and subclass in WEAPON_SKILLS:
                        skill_id, skill_name = WEAPON_SKILLS[subclass]
                        correct_inv_type = SKILL_TO_INVTYPE.get(skill_id, correct_inv_type)

                    # Ammo (class 6) - always invType 24
                    elif item_class == 6:
                        correct_inv_type = 24

                    # Only fix if invType is missing/wrong (NULL, -1, or different from item_template)
                    # Note: inv_type can be None (NULL), -1 (missing), or any other value
                    # We normalize None to -1 for comparison
                    current_inv_type = inv_type if inv_type is not None else -1

                    # Only add to fixes if the database value differs from the correct value
                    if current_inv_type != correct_inv_type:
                        invtype_fixes.append({
                            'outfit_id': outfit_id,
                            'race': race_name,
                            'class': class_name,
                            'gender': gender_name,
                            'slot': slot_idx,
                            'item_id': item_id,
                            'item_name': item_name,
                            'inv_type': correct_inv_type
                        })

        print(f"Found {len(invtype_fixes)} items with missing invType")
        print()

        # Scan for duplicates (always run this)
        print("=" * 80)
        print("SCANNING FOR DUPLICATE WEAPONS/AMMO")
        print("=" * 80)
        print()

        duplicate_cleanups = []
        dbc_cursor.execute("""
            SELECT o.ID, o.race, o.class, o.gender,
                   o.itemId1, o.itemId2, o.itemId3, o.itemId4, o.itemId5,
                   o.itemId6, o.itemId7, o.itemId8, o.itemId9, o.itemId10,
                   o.itemId11, o.itemId12, o.itemId13, o.itemId14, o.itemId15,
                   o.itemId16, o.itemId17, o.itemId18, o.itemId19, o.itemId20,
                   o.itemId21, o.itemId22, o.itemId23, o.itemId24,
                   o.invType1, o.invType2, o.invType3, o.invType4, o.invType5,
                   o.invType6, o.invType7, o.invType8, o.invType9, o.invType10,
                   o.invType11, o.invType12, o.invType13, o.invType14, o.invType15,
                   o.invType16, o.invType17, o.invType18, o.invType19, o.invType20,
                   o.invType21, o.invType22, o.invType23, o.invType24
            FROM charstartoutfit o
            WHERE o.race BETWEEN 1 AND 12 AND o.class BETWEEN 1 AND 11
        """)
        current_outfits = dbc_cursor.fetchall()

        for outfit in current_outfits:
            outfit_id = outfit[0]
            race_id = outfit[1]
            class_id = outfit[2]
            gender = outfit[3]
            item_ids = outfit[4:28]  # itemId1-24 (24 slots)
            inv_types = outfit[28:52]  # invType1-24 (24 slots)

            race_name = RACE_NAMES.get(race_id, f"Race{race_id}")
            class_name = CLASS_NAMES.get(class_id, f"Class{class_id}")
            gender_name = "Male" if gender == 0 else "Female"

            # Track weapons and ammo by type to detect duplicates
            weapon_slots_by_type = {}
            weapon_inv_types = [13, 15, 17, 21, 22, 26]
            ammo_inv_type = 24

            # Also track melee/ranged weapons across types for multi-weapon detection
            melee_weapon_slots = []  # [(slot, skill_id, skill_name), ...]
            ranged_weapon_slots = []  # [(slot, skill_id, skill_name), ...]
            melee_skills = [43, 44, 54, 55, 136, 160, 172, 173]  # All melee weapon types
            ranged_skills = [45, 46, 226, 176, 228]  # Bows, Guns, Crossbows, Thrown, Wands

            for slot_idx, (item_id, inv_type) in enumerate(zip(item_ids, inv_types), 1):
                if not item_id or item_id <= 0 or item_id == 6948:
                    continue

                # Check weapons
                if inv_type in weapon_inv_types:
                    acore_cursor.execute(
                        "SELECT subclass FROM item_template WHERE entry = %s AND class = 2",
                        (item_id,)
                    )
                    result = acore_cursor.fetchone()
                    if result and result[0] in WEAPON_SKILLS:
                        skill_id, skill_name = WEAPON_SKILLS[result[0]]
                        type_key = f"weapon_{skill_id}"
                        if type_key not in weapon_slots_by_type:
                            weapon_slots_by_type[type_key] = {'name': skill_name, 'slots': []}
                        weapon_slots_by_type[type_key]['slots'].append(slot_idx)

                        # Track melee vs ranged
                        if skill_id in melee_skills:
                            melee_weapon_slots.append((slot_idx, skill_id, skill_name))
                        elif skill_id in ranged_skills:
                            ranged_weapon_slots.append((slot_idx, skill_id, skill_name))

                # Check ammo
                elif inv_type == ammo_inv_type:
                    acore_cursor.execute(
                        "SELECT subclass, name FROM item_template WHERE entry = %s AND class = 6",
                        (item_id,)
                    )
                    result = acore_cursor.fetchone()
                    if result:
                        ammo_subclass = result[0]
                        ammo_name = result[1]
                        type_key = f"ammo_{ammo_subclass}"
                        if type_key not in weapon_slots_by_type:
                            weapon_slots_by_type[type_key] = {'name': ammo_name, 'slots': []}
                        weapon_slots_by_type[type_key]['slots'].append(slot_idx)

            # Find duplicates of same weapon type
            for type_key, type_data in weapon_slots_by_type.items():
                slots = type_data['slots']
                if len(slots) > 1:
                    item_name = type_data['name']
                    for slot_idx in slots[1:]:
                        duplicate_cleanups.append({
                            'outfit_id': outfit_id,
                            'race': race_name,
                            'class': class_name,
                            'gender': gender_name,
                            'slot': slot_idx,
                            'item_name': item_name,
                            'reason': 'Duplicate same weapon type'
                        })

            # NEW: Detect multiple different melee weapons (e.g., axe + dagger)
            if len(melee_weapon_slots) > 1:
                # Keep first melee weapon, remove the rest
                for slot_idx, skill_id, skill_name in melee_weapon_slots[1:]:
                    duplicate_cleanups.append({
                        'outfit_id': outfit_id,
                        'race': race_name,
                        'class': class_name,
                        'gender': gender_name,
                        'slot': slot_idx,
                        'item_name': skill_name,
                        'reason': f'Multiple melee weapons (keeping first, removing extra {skill_name})'
                    })

            # NEW: Detect multiple different ranged weapons (e.g., bow + gun)
            if len(ranged_weapon_slots) > 1:
                # Keep first ranged weapon, remove the rest
                for slot_idx, skill_id, skill_name in ranged_weapon_slots[1:]:
                    duplicate_cleanups.append({
                        'outfit_id': outfit_id,
                        'race': race_name,
                        'class': class_name,
                        'gender': gender_name,
                        'slot': slot_idx,
                        'item_name': skill_name,
                        'reason': f'Multiple ranged weapons (keeping first, removing extra {skill_name})'
                    })

        print(f"Found {len(duplicate_cleanups)} duplicate weapons/ammo to clean up")
        print()

        # Generate SQL fixes
        if mismatches or weapon_additions or duplicate_cleanups or invtype_fixes:
            print("=" * 80)
            print("GENERATING STARTING WEAPONS SQL")
            print("=" * 80)
            print()

            # Build SQL lines
            sql_lines = []

            # SECTION 1: Fix missing invType values (must run first!)
            if invtype_fixes:
                sql_lines.append("-- " + "=" * 76)
                sql_lines.append("-- FIX MISSING INVTYPE VALUES")
                sql_lines.append("-- " + "=" * 76)
                sql_lines.append("-- This section fixes items that have itemId but missing invType.")
                sql_lines.append("-- Must run BEFORE duplicate detection can work properly.")
                sql_lines.append("")

                for fix in invtype_fixes:
                    sql_lines.append(f"-- {fix['race']} {fix['class']} ({fix['gender']}): Set invType for {fix['item_name']} (item {fix['item_id']})")
                    sql_lines.append(f"UPDATE `dbc`.`charstartoutfit` SET `invType{fix['slot']}` = {fix['inv_type']} WHERE `ID` = {fix['outfit_id']};")

                sql_lines.append("")
                sql_lines.append(f"-- Fixed invType for {len(invtype_fixes)} items")
                sql_lines.append("")
                sql_lines.append("")

            # Combine and sort all fixes by outfit_id
            all_fixes = []

            for mismatch in mismatches:
                all_fixes.append({
                    'outfit_id': mismatch['outfit_id'],
                    'race': mismatch['race'],
                    'class': mismatch['class'],
                    'gender': mismatch['gender'],
                    'slot': mismatch['slot'],
                    'type': 'replace',
                    'data': mismatch
                })

            for addition in weapon_additions:
                all_fixes.append({
                    'outfit_id': addition['outfit_id'],
                    'race': addition['race'],
                    'class': addition['class'],
                    'gender': addition['gender'],
                    'slot': addition['slot'],
                    'type': 'add',
                    'data': addition
                })

            # Sort by outfit_id, then by slot
            all_fixes.sort(key=lambda x: (x['outfit_id'], x['slot']))

            # SECTION 2: Weapon replacements and additions
            warning_count = 0

            for fix in all_fixes:
                slot = fix['slot']
                outfit_id = fix['outfit_id']

                if fix['type'] == 'replace':
                    mismatch = fix['data']
                    new_item = mismatch['replacement_item'] if mismatch['replacement_item'] else -1

                    if new_item == -1:
                        warning_count += 1
                        sql_lines.append(f"-- WARNING: {mismatch['race']} {mismatch['class']} ({mismatch['gender']}): No suitable replacement for {mismatch['current_type']}")
                        sql_lines.append(f"-- Available skills: {mismatch['available_skills']}")
                        sql_lines.append(f"-- UPDATE `dbc`.`charstartoutfit` SET `itemId{slot}` = -1 WHERE `ID` = {outfit_id};")
                    else:
                        skill_name = next((name for sid, name in WEAPON_SKILLS.values() if sid == mismatch['replacement_skill']), "Unknown")
                        sql_lines.append(f"-- {mismatch['race']} {mismatch['class']} ({mismatch['gender']}): Replace {mismatch['current_type']} with {skill_name}")
                        sql_lines.append(f"UPDATE `dbc`.`charstartoutfit` SET `itemId{slot}` = {new_item} WHERE `ID` = {outfit_id};")
                    sql_lines.append("")
                else:  # add
                    addition = fix['data']
                    # Handle ammo (add_skill is None for ammo)
                    if addition.get('is_ammo'):
                        sql_lines.append(f"-- {addition['race']} {addition['class']} ({addition['gender']}): Add {addition['reason']}")
                        sql_lines.append(f"UPDATE `dbc`.`charstartoutfit` SET `itemId{slot}` = {addition['add_item']}, `invType{slot}` = 24 WHERE `ID` = {outfit_id};")
                    else:
                        skill_name = next((name for sid, name in WEAPON_SKILLS.values() if sid == addition['add_skill']), "Unknown")
                        inv_type = SKILL_TO_INVTYPE.get(addition['add_skill'], 13)  # Default to one-hand if not found
                        sql_lines.append(f"-- {addition['race']} {addition['class']} ({addition['gender']}): Add {skill_name} ({addition['reason']})")
                        sql_lines.append(f"UPDATE `dbc`.`charstartoutfit` SET `itemId{slot}` = {addition['add_item']}, `invType{slot}` = {inv_type} WHERE `ID` = {outfit_id};")
                    sql_lines.append("")

            # SECTION 3: Cleanup duplicates
            if duplicate_cleanups:
                sql_lines.append("-- " + "=" * 76)
                sql_lines.append("-- CLEANUP: REMOVE DUPLICATE WEAPONS/AMMO")
                sql_lines.append("-- " + "=" * 76)
                sql_lines.append("-- This section removes duplicate weapons/ammo from the database.")
                sql_lines.append("-- Runs AFTER invType fixes, so duplicates can be properly detected.")
                sql_lines.append("")

                for cleanup in duplicate_cleanups:
                    reason = cleanup.get('reason', 'Duplicate weapon/ammo')
                    sql_lines.append(f"-- {cleanup['race']} {cleanup['class']} ({cleanup['gender']}): {reason}")
                    sql_lines.append(f"UPDATE `dbc`.`charstartoutfit` SET `itemId{cleanup['slot']}` = 0, `invType{cleanup['slot']}` = 0 WHERE `ID` = {cleanup['outfit_id']};")

                sql_lines.append("")
                sql_lines.append(f"-- Cleaned up {len(duplicate_cleanups)} duplicate weapons/ammo")
                sql_lines.append("")

            # Write to file (same directory as script)
            script_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(script_dir, 'starting_weapons.sql')
            with open(output_path, 'w') as f:
                f.write('\n'.join(sql_lines))

            weapons_count = len([a for a in weapon_additions if not a.get('is_ammo')])
            ammo_count = len([a for a in weapon_additions if a.get('is_ammo')])
            print(f"✓ Generated {output_path}")
            print(f"  {len(mismatches)} weapon replacements")
            print(f"  {weapons_count} weapon additions")
            print(f"  {ammo_count} ammo additions")
            print(f"  {len(duplicate_cleanups)} duplicate weapons/ammo cleaned up")
            print(f"  {warning_count} warnings")
        else:
            print("✓ NO FIXES NEEDED - All starting gear is valid!")

    dbc_cursor.close()
    acore_cursor.close()
    dbc_conn.close()
    acore_conn.close()

    return 1 if (mismatches or weapon_additions) else 0

if __name__ == "__main__":
    sys.exit(main())
